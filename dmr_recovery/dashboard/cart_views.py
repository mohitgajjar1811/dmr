from io import BytesIO
import csv
from decimal import Decimal
from datetime import date
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.template.loader import render_to_string
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone

from cart.models import Cart, CartItem
from core.utils import send_custom_email

def is_admin(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)

def generate_quotation_excel(cart):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Blue
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Header Info
    ws['A1'] = "QUOTATION"
    ws['A1'].font = Font(size=20, bold=True)
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A3'] = "Customer Details:"
    ws['A3'].font = Font(bold=True)
    ws['A4'] = f"Name: {cart.name}"
    ws['A5'] = f"Company: {cart.company_name or 'N/A'}"
    ws['A6'] = f"Email: {cart.email}"
    ws['A7'] = f"Phone: {cart.phone}"
    ws['A8'] = f"Location: {cart.city or ''}, {cart.state or ''}"
    
    # Table Header
    headers = ['SR#', 'Product Name', 'Quantity', 'Rate (₹)', 'Total (₹)']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        
    # Table Content
    row_num = 11
    subtotal = 0
    for idx, item in enumerate(cart.items.all(), 1):
        line_total = Decimal(str(item.quantity)) * item.rate
        subtotal += line_total
        
        ws.cell(row=row_num, column=1, value=idx).border = border
        ws.cell(row=row_num, column=2, value=item.product.name).border = border
        ws.cell(row=row_num, column=3, value=item.quantity).border = border
        ws.cell(row=row_num, column=4, value=item.rate).border = border
        ws.cell(row=row_num, column=5, value=line_total).border = border
        
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='center')
        
        row_num += 1
        
    # Totals
    row_num += 1
    ws.cell(row=row_num, column=4, value="Subtotal:").font = Font(bold=True)
    ws.cell(row=row_num, column=5, value=subtotal).font = Font(bold=True)
    
    row_num += 1
    state = (cart.state or '').strip().lower()
    is_maharashtra = 'maharashtra' in state
    
    if is_maharashtra:
        cgst = subtotal * Decimal('0.09')
        sgst = subtotal * Decimal('0.09')
        ws.cell(row=row_num, column=4, value="CGST (9%):")
        ws.cell(row=row_num, column=5, value=cgst)
        row_num += 1
        ws.cell(row=row_num, column=4, value="SGST (9%):")
        ws.cell(row=row_num, column=5, value=sgst)
        grand_total = subtotal + cgst + sgst
    else:
        igst = subtotal * Decimal('0.18')
        ws.cell(row=row_num, column=4, value="IGST (18%):")
        ws.cell(row=row_num, column=5, value=igst)
        grand_total = subtotal + igst
        
    row_num += 1
    ws.cell(row=row_num, column=4, value="Grand Total:").font = Font(bold=True, size=12)
    ws.cell(row=row_num, column=5, value=grand_total).font = Font(bold=True, size=12)
    
    # Column Widths
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    ws['A20'] = "This is a computer generated quotation."
    ws['A20'].font = Font(italic=True, size=8)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

@login_required
@user_passes_test(is_admin, login_url='/dashboard/login/')
def user_carts_list(request):
    # Only get submitted carts
    carts = Cart.objects.filter(is_submitted=True).order_by('-updated_at')
    
    # Filtering logic
    q = request.GET.get('q')
    status = request.GET.get('status')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if q:
        carts = carts.filter(
            Q(name__icontains=q) |
            Q(email__icontains=q) |
            Q(phone__icontains=q) |
            Q(company_name__icontains=q) |
            Q(city__icontains=q) |
            Q(state__icontains=q)
        )
    
    if status == 'unread':
        carts = carts.filter(is_read=False)
    elif status == 'read':
        carts = carts.filter(is_read=True, is_replied=False)
    elif status == 'replied':
        carts = carts.filter(is_replied=True)

    if date_from:
        carts = carts.filter(updated_at__date__gte=date_from)
    if date_to:
        carts = carts.filter(updated_at__date__lte=date_to)
    
    context = {
        'carts': carts,
        'page_title': 'User Carts',
        'q': q,
        'status': status,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'dashboard/carts/list.html', context)

@login_required
@user_passes_test(is_admin, login_url='/dashboard/login/')
def user_cart_detail(request, cart_id):
    cart = get_object_or_404(Cart, id=cart_id, is_submitted=True)
    
    # Mark as read
    if not cart.is_read:
        cart.is_read = True
        cart.save()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        # Always update rates if they are in the POST
        for key, value in request.POST.items():
            if key.startswith('rate_'):
                item_id = key.replace('rate_', '')
                try:
                    item = cart.items.get(id=item_id)
                    item.rate = float(value or 0)
                    item.save()
                except (ValueError, CartItem.DoesNotExist):
                    pass
        
        if action == 'delete':
            cart.delete()
            messages.success(request, 'Quote Request deleted successfully.')
            return redirect('dashboard:dashboard_carts_list')
        elif action == 'reply':
            reply_message = request.POST.get('reply_message')
            if reply_message:
                cart.is_replied = True
                cart.reply_message = reply_message
                cart.replied_at = timezone.now()
                cart.save()
                
                # Generate Quotation Excel
                try:
                    excel_content = generate_quotation_excel(cart)
                    filename = f"Quotation_{cart.id}_{date.today().strftime('%Y%m%d')}.xlsx"
                    attachments = [(filename, excel_content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')]
                    
                    subject = f"Your Quote Quotation - Ref: #{cart.id}"
                    html_content = render_to_string('dashboard/emails/cart_reply.html', {'cart': cart})
                    text_content = f"Hello {cart.name},\n\nPlease find the attached quotation for your request.\n\nOur Message:\n{cart.reply_message}\n\nThank you for choosing Virja."
                    
                    send_custom_email(subject, text_content, [cart.email], html_message=html_content, attachments=attachments)
                    messages.success(request, f"Reply sent successfully to {cart.email} with Quotation Attachment!")
                except Exception as e:
                    messages.warning(request, f"Reply saved, but failed to generate/send attachment. Error: {str(e)}")
                    
            return redirect('dashboard:dashboard_cart_detail', cart_id=cart.id)
            
    context = {
        'cart': cart,
        'page_title': f'Cart Details',
    }
    return render(request, 'dashboard/carts/detail.html', context)

@login_required
@user_passes_test(is_admin, login_url='/dashboard/login/')
def user_cart_delete(request, cart_id):
    cart = get_object_or_404(Cart, id=cart_id, is_submitted=True)
    if request.method == 'POST':
        cart.delete()
        messages.success(request, "Quote requested deleted successfully.")
        return redirect('dashboard:dashboard_carts_list')
    return render(request, 'dashboard/catalog/category_confirm_delete.html', {
        'item': cart,
        'type': 'Quote Request',
        'cancel_url': 'dashboard:dashboard_carts_list'
    })

@login_required
@user_passes_test(is_admin, login_url='/dashboard/login/')
def user_carts_export(request):
    carts = Cart.objects.filter(is_submitted=True).order_by('-updated_at')
    
    # Apply same filters for export
    q = request.GET.get('q')
    status = request.GET.get('status')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if q:
        carts = carts.filter(
            Q(name__icontains=q) |
            Q(email__icontains=q) |
            Q(phone__icontains=q) |
            Q(company_name__icontains=q) |
            Q(city__icontains=q) |
            Q(state__icontains=q)
        )
    if status == 'unread':
        carts = carts.filter(is_read=False)
    elif status == 'read':
        carts = carts.filter(is_read=True, is_replied=False)
    elif status == 'replied':
        carts = carts.filter(is_replied=True)
    if date_from:
        carts = carts.filter(updated_at__date__gte=date_from)
    if date_to:
        carts = carts.filter(updated_at__date__lte=date_to)

    today = date.today().strftime('%d-%m-%Y')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="customer_details_{today}.csv"'
    
    writer = csv.writer(response)
    # Only customer details as requested
    writer.writerow(['Date', 'Name', 'Email', 'Phone', 'Company', 'State', 'City'])
    
    for cart in carts:
        writer.writerow([
            cart.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
            cart.name,
            cart.email,
            cart.phone,
            cart.company_name,
            cart.state,
            cart.city
        ])
        
    return response
